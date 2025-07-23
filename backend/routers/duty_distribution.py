from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from typing import List, Dict, Any, Optional
from database import get_db
from models.models import Department, Employee, DutyType, DutyRecord, EmployeeDutyType, DepartmentDutyDay
from pydantic import BaseModel
from datetime import datetime, date, timedelta
import calendar
import random
from fastapi.responses import StreamingResponse
import io
import logging
import traceback

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Настройка логирования
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

router = APIRouter()

class DutyDistributionRequest(BaseModel):
    start_date: str
    end_date: str
    department_id: Optional[int] = None
    structure_id: Optional[int] = None

class DutyDistributionResponse(BaseModel):
    department_id: int
    department_name: str
    duties: List[Dict[str, Any]]

@router.post("/generate", response_model=List[DutyDistributionResponse])
async def generate_duty_distribution(
    request: DutyDistributionRequest, 
    db: AsyncSession = Depends(get_db)
):
    """Генерировать распределение нарядов на выбранный период для конкретного подразделения"""
    
    logger.debug(f"Начало генерации нарядов")
    logger.debug(f"Параметры: start_date={request.start_date}, end_date={request.end_date}, department_id={request.department_id}")
    
    # Парсим даты
    start_date = datetime.strptime(request.start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(request.end_date, "%Y-%m-%d").date()
    
    # Получаем все типы нарядов
    duty_types_result = await db.execute(select(DutyType))
    duty_types = duty_types_result.scalars().all()
    
    logger.debug(f"Найдено типов нарядов: {len(duty_types)}")
    for dt in duty_types:
        logger.debug(f"Тип наряда {dt.id}: {dt.name} (категория: {dt.duty_category})")
    
    # Формируем запрос для получения сотрудников
    employee_query = (
            select(Employee, EmployeeDutyType, DutyType)
            .join(EmployeeDutyType, Employee.id == EmployeeDutyType.employee_id)
            .join(DutyType, EmployeeDutyType.duty_type_id == DutyType.id)
            .where(Employee.is_active == True)
            .where(EmployeeDutyType.is_active == True)
        )
        
    # Если указано конкретное подразделение
    if request.department_id:
        employee_query = employee_query.where(Employee.department_id == request.department_id)
    # Если указана структура (но не подразделение)
    elif request.structure_id:
        # Получаем все подразделения структуры
        subdepts_result = await db.execute(
            select(Department.id).where(Department.parent_id == request.structure_id)
        )
        subdept_ids = [row[0] for row in subdepts_result.all()]
        if subdept_ids:
            employee_query = employee_query.where(Employee.department_id.in_(subdept_ids))
    
    employees_result = await db.execute(employee_query)
    employees_data = employees_result.all()
    
    # Группируем сотрудников по типам нарядов
    duty_types_employees = {}
    for emp, emp_duty, duty_type in employees_data:
        if duty_type.id not in duty_types_employees:
            duty_types_employees[duty_type.id] = {
                'duty_type': duty_type,
                'employees': []
            }
        duty_types_employees[duty_type.id]['employees'].append(emp)
        
    # Словарь для отслеживания занятости сотрудников по дням и типам нарядов
    # Формат: {employee_id: {date_key: {duty_type_id, ...}}}
    employee_busy_dates = {}
    
    # Собираем все наряды
    all_duties = []
    
    # Генерируем наряды по дням, а не по типам
    current_date = start_date
    while current_date <= end_date:
        duty_date = current_date
        date_key = duty_date.isoformat()
        
        # Для каждого типа наряда в этот день
        for duty_type_id, data in duty_types_employees.items():
            duty_type = data['duty_type']
            employees = data['employees']
            
            # Проверяем, нужно ли назначать академический наряд в этот день
            if duty_type.duty_category == "academic":
                logger.debug(f"Проверяем академический наряд {duty_type.name} для даты {duty_date}")
                
                # Проверяем, есть ли запись в календаре для этого типа наряда и подразделения
                calendar_query = (
                    select(DepartmentDutyDay)
                    .where(DepartmentDutyDay.duty_type_id == duty_type_id)
                    .where(DepartmentDutyDay.duty_date == duty_date)
                )
                
                # Если указано конкретное подразделение
                if request.department_id:
                    calendar_query = calendar_query.where(DepartmentDutyDay.department_id == request.department_id)
                    logger.debug(f"Фильтруем по подразделению {request.department_id}")
                # Если указана структура, проверяем все подразделения структуры
                elif request.structure_id:
                    subdepts_result = await db.execute(
                        select(Department.id).where(Department.parent_id == request.structure_id)
                    )
                    subdept_ids = [row[0] for row in subdepts_result.all()]
                    if subdept_ids:
                        calendar_query = calendar_query.where(DepartmentDutyDay.department_id.in_(subdept_ids))
                        logger.debug(f"Фильтруем по подразделениям структуры {subdept_ids}")
                    else:
                        # Если нет подразделений в структуре, пропускаем
                        logger.debug(f"Нет подразделений в структуре {request.structure_id}")
                        continue
                
                calendar_check_result = await db.execute(calendar_query)
                calendar_records = calendar_check_result.scalars().all()
                
                logger.debug(f"Найдено записей в календаре: {len(calendar_records)}")
                
                # Если нет записей в календаре, пропускаем этот день для академического наряда
                if not calendar_records:
                    logger.debug(f"Нет записей в календаре, пропускаем")
                    continue
                else:
                    logger.debug(f"Обрабатываем записи календаря")
                    # Для каждого дня из календаря выбираем сотрудников подразделения
                    for calendar_record in calendar_records:
                        department_id = calendar_record.department_id
                        logger.debug(f"Обрабатываем подразделение {department_id}")
                        
                        # Получаем сотрудников подразделения, которые могут заступать в этот тип наряда
                        department_employees_result = await db.execute(
                            select(Employee)
                            .join(EmployeeDutyType, Employee.id == EmployeeDutyType.employee_id)
                            .where(Employee.department_id == department_id)
                            .where(EmployeeDutyType.duty_type_id == duty_type_id)
                            .where(EmployeeDutyType.is_active == True)
                            .where(Employee.is_active == True)
                        )
                        department_employees = department_employees_result.scalars().all()
                        
                        logger.debug(f"Найдено сотрудников в подразделении {department_id}: {len(department_employees)}")
                        
                        if not department_employees:
                            logger.debug(f"Нет сотрудников в подразделении {department_id}")
                            continue
                        
                        # Выбираем сотрудников для наряда
                        selected_employees = await select_employees_for_duty(
                            department_employees, duty_date, duty_type.people_per_day, 
                            employee_busy_dates, duty_type_id, db, start_date, end_date
                        )
                        
                        logger.debug(f"Выбрано сотрудников для наряда: {len(selected_employees)}")
                        
                        # Назначаем выбранных сотрудников
                        for selected_employee in selected_employees:
                            # Блокируем сотрудника на все дни длительности наряда
                            for day_offset in range(duty_type.days_duration):
                                duty_day = duty_date + timedelta(days=day_offset)
                                duty_day_key = duty_day.isoformat()
                                
                                # Проверяем, не выходит ли день за пределы периода
                                if duty_day > end_date:
                                    break
                                
                                if selected_employee.id not in employee_busy_dates:
                                    employee_busy_dates[selected_employee.id] = {}
                                if duty_day_key not in employee_busy_dates[selected_employee.id]:
                                    employee_busy_dates[selected_employee.id][duty_day_key] = set()
                                employee_busy_dates[selected_employee.id][duty_day_key].add(duty_type_id)
                                
                                # Добавляем наряд только для первого дня (начало наряда)
                                if day_offset == 0:
                                    all_duties.append({
                                        'date': duty_day_key,
                                        'employee_id': selected_employee.id,
                                        'employee_name': f"{selected_employee.last_name} {selected_employee.first_name}",
                                        'duty_type_id': duty_type.id,
                                        'duty_type_name': duty_type.name,
                                        'people_per_day': duty_type.people_per_day,
                                        'days_duration': duty_type.days_duration
                                    })
                            logger.debug(f"Добавлен наряд для {selected_employee.last_name} {selected_employee.first_name}")
                    
                    # Пропускаем автоматический выбор сотрудников для академических нарядов
                    continue
            
            # Для каждого типа наряда должно заступить столько людей, сколько указано в people_per_day
            people_needed = duty_type.people_per_day
            
            # Выбираем сотрудников для наряда
            selected_employees = await select_employees_for_duty(
                employees, duty_date, people_needed, employee_busy_dates, duty_type_id, db, start_date, end_date
            )
            
            # Назначаем выбранных сотрудников
            for selected_employee in selected_employees:
                # Блокируем сотрудника на все дни длительности наряда
                for day_offset in range(duty_type.days_duration):
                    duty_day = duty_date + timedelta(days=day_offset)
                    duty_day_key = duty_day.isoformat()
                    
                    # Проверяем, не выходит ли день за пределы периода
                    if duty_day > end_date:
                        break
                    
                    if selected_employee.id not in employee_busy_dates:
                        employee_busy_dates[selected_employee.id] = {}
                    if duty_day_key not in employee_busy_dates[selected_employee.id]:
                        employee_busy_dates[selected_employee.id][duty_day_key] = set()
                    employee_busy_dates[selected_employee.id][duty_day_key].add(duty_type_id)
                    
                    # Добавляем наряд только для первого дня (начало наряда)
                    if day_offset == 0:
                        all_duties.append({
                            'date': duty_day_key,
                        'employee_id': selected_employee.id,
                        'employee_name': f"{selected_employee.last_name} {selected_employee.first_name}",
                        'duty_type_id': duty_type.id,
                            'duty_type_name': duty_type.name,
                            'people_per_day': duty_type.people_per_day,
                            'days_duration': duty_type.days_duration
                        })
        
        current_date += timedelta(days=1)
    
    # Группируем по подразделениям для ответа
    dept_result = await db.execute(select(Department))
    departments = dept_result.scalars().all()
    
    distribution = []
    for dept in departments:
        # Получаем сотрудников подразделения для проверки
        dept_employees_result = await db.execute(
            select(Employee.id, Employee.duty_count).where(Employee.department_id == dept.id)
        )
        dept_employees_data = dept_employees_result.all()
        dept_employee_ids = [row[0] for row in dept_employees_data]
        dept_employee_duty_counts = {row[0]: row[1] or 0 for row in dept_employees_data}
        
        # Фильтруем наряды для этого подразделения
        dept_duties = [d for d in all_duties if d['employee_id'] in dept_employee_ids]
        
        # Добавляем duty_count к каждому наряду
        for duty in dept_duties:
            duty['duty_count'] = dept_employee_duty_counts.get(duty['employee_id'], 0)
        
        if dept_duties:
            distribution.append({
                'department_id': dept.id,
                'department_name': dept.name,
                'duties': dept_duties
            })
    
    # Сохраняем все наряды в базу
    for duty in all_duties:
        # Получаем длительность наряда
        duty_duration = duty.get('days_duration', 1)
        start_date_duty = date.fromisoformat(duty['date'])
        
        # Создаем записи для всех дней длительности наряда
        for day_offset in range(duty_duration):
            duty_day = start_date_duty + timedelta(days=day_offset)
            
            # Проверяем, не выходит ли день за пределы периода
            if duty_day > end_date:
                break
            
            # Проверяем, не существует ли уже такая запись
            existing_record = await db.execute(
                select(DutyRecord)
                .where(DutyRecord.employee_id == duty['employee_id'])
                .where(DutyRecord.duty_type_id == duty['duty_type_id'])
                .where(DutyRecord.duty_date == duty_day)
            )
            existing_record = existing_record.scalar_one_or_none()
            
            if not existing_record:
                # Создаем новую запись
                db.add(DutyRecord(
                    employee_id=duty['employee_id'],
                    duty_type_id=duty['duty_type_id'],
                    duty_date=duty_day
                ))
    
    await db.commit()
    return distribution

async def select_employees_for_duty(
    employees: List[Employee],
    duty_date: date,
    people_needed: int,
    employee_busy_dates: Dict[int, Dict[str, set]],
    duty_type_id: int,
    db: AsyncSession,
    start_date: date,
    end_date: date
) -> List[Employee]:
    """Выбирает сотрудников для наряда с учетом количества нарядов за период и ограничения интервалов между нарядами"""
    date_key = duty_date.isoformat()
    
    # Получаем количество нарядов каждого сотрудника за выбранный период
    employee_duty_counts = {}
    employee_last_duty_dates_by_type = {}
    employee_last_duty_info = {}  # Информация о последнем наряде (дата и длительность)
    
    for employee in employees:
        # Количество нарядов за выбранный период (из базы)
        count_result = await db.execute(
            select(func.count(DutyRecord.id))
            .where(DutyRecord.employee_id == employee.id)
            .where(DutyRecord.duty_date >= start_date)
            .where(DutyRecord.duty_date <= end_date)
        )
        count_in_db = count_result.scalar() or 0
        
        # Количество нарядов в памяти (уже назначенных в этой сессии)
        count_in_memory = 0
        if employee.id in employee_busy_dates:
            for date_str in employee_busy_dates[employee.id]:
                count_in_memory += len(employee_busy_dates[employee.id][date_str])
        
        # Общее количество нарядов = наряды в выбранном периоде + duty_count
        total_duty_count = count_in_db + count_in_memory + (employee.duty_count or 0)
        employee_duty_counts[employee.id] = total_duty_count
        
        # Дата последнего наряда этого типа
        last_duty_result = await db.execute(
            select(DutyRecord.duty_date)
            .where(DutyRecord.employee_id == employee.id)
            .where(DutyRecord.duty_type_id == duty_type_id)
            .order_by(DutyRecord.duty_date.desc())
            .limit(1)
        )
        last_duty_date = last_duty_result.scalar_one_or_none()
        
        # Учитываем наряды в памяти
        all_dates = set()
        if last_duty_date:
            all_dates.add(last_duty_date)
        if employee.id in employee_busy_dates:
            for date_str in employee_busy_dates[employee.id]:
                if duty_type_id in employee_busy_dates[employee.id][date_str]:
                    all_dates.add(date.fromisoformat(date_str))
        
        last_duty_date = max(all_dates) if all_dates else None
        employee_last_duty_dates_by_type[(employee.id, duty_type_id)] = last_duty_date
        
        # Получаем информацию о последнем наряде любого типа (дата и длительность)
        last_any_duty_result = await db.execute(
            select(DutyRecord.duty_date, DutyType.days_duration)
            .join(DutyType, DutyRecord.duty_type_id == DutyType.id)
            .where(DutyRecord.employee_id == employee.id)
            .order_by(DutyRecord.duty_date.desc())
            .limit(1)
        )
        last_any_duty_data = last_any_duty_result.first()
        
        # Учитываем наряды в памяти для любого типа
        all_any_dates = set()
        last_duty_duration = 1  # По умолчанию 1 день
        
        if last_any_duty_data:
            all_any_dates.add(last_any_duty_data[0])
            last_duty_duration = last_any_duty_data[1] or 1
        
        if employee.id in employee_busy_dates:
            for date_str in employee_busy_dates[employee.id]:
                all_any_dates.add(date.fromisoformat(date_str))
                # Для нарядов в памяти считаем длительность 1 день (так как это текущая сессия)
                last_duty_duration = 1
        
        last_any_duty_date = max(all_any_dates) if all_any_dates else None
        employee_last_duty_info[employee.id] = {
            'date': last_any_duty_date,
            'duration': last_duty_duration
        }
    
    # Фильтруем доступных сотрудников
    available_employees = []
    for employee in employees:
        # Проверяем, не занят ли сотрудник в этот тип наряда в эту дату
        if (employee.id in employee_busy_dates and 
            date_key in employee_busy_dates[employee.id] and 
            duty_type_id in employee_busy_dates[employee.id][date_key]):
            continue
        
        # Проверяем ограничение интервалов между нарядами в зависимости от длительности предыдущего наряда
        last_duty_info = employee_last_duty_info[employee.id]
        if last_duty_info['date'] is not None:
            days_since_last_duty = (duty_date - last_duty_info['date']).days
            last_duty_duration = last_duty_info['duration']
            
            # Вычисляем минимальный интервал в зависимости от длительности предыдущего наряда
            # 1 день = следующий день (интервал 1 день)
            # 2 дня = через день (интервал 2 дня)  
            # 3 дня = через 2 дня (интервал 3 дня)
            min_interval = last_duty_duration
            
            if days_since_last_duty < min_interval:
                continue
        
        available_employees.append(employee)
    
    if not available_employees:
        return []
    
    # Сортируем по общему количеству нарядов (приоритет тем, у кого меньше нарядов)
    available_employees.sort(key=lambda emp: employee_duty_counts[emp.id])
    
    # Выбираем сотрудников с наименьшим общим количеством нарядов
    selected_employees = []
    current_count = employee_duty_counts[available_employees[0].id]
    
    # Сначала берем всех сотрудников с минимальным количеством нарядов
    candidates = [emp for emp in available_employees if employee_duty_counts[emp.id] == current_count]
    
    # Если кандидатов больше, чем нужно, выбираем тех, кто давно не был в этом типе наряда
    if len(candidates) > people_needed:
        candidates.sort(key=lambda emp: employee_last_duty_dates_by_type[(emp.id, duty_type_id)] or date.min)
    
    selected_employees = candidates[:people_needed]
    
    # Если нужно больше сотрудников, берем следующих по количеству нарядов
    if len(selected_employees) < people_needed:
        remaining_needed = people_needed - len(selected_employees)
        
        # Берем следующих по количеству нарядов
        next_candidates = [emp for emp in available_employees if employee_duty_counts[emp.id] > current_count]
        next_candidates.sort(key=lambda emp: employee_duty_counts[emp.id])
        
        # Добавляем сотрудников с следующим количеством нарядов
        for next_count in sorted(set(employee_duty_counts[emp.id] for emp in next_candidates)):
            if len(selected_employees) >= people_needed:
                break
            
            same_count_candidates = [emp for emp in next_candidates if employee_duty_counts[emp.id] == next_count]
            same_count_candidates.sort(key=lambda emp: employee_last_duty_dates_by_type[(emp.id, duty_type_id)] or date.min)
            
            needed_from_this_count = min(remaining_needed - len(selected_employees), len(same_count_candidates))
            selected_employees.extend(same_count_candidates[:needed_from_this_count])
    
    return selected_employees

@router.get("/department/{department_id}")
async def get_duty_distribution_by_department(
    department_id: int,
    start_date: str = Query(None, description="Начальная дата (YYYY-MM-DD)"),
    end_date: str = Query(None, description="Конечная дата (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db)
):
    """Получить распределение нарядов для конкретного подразделения за выбранный период"""
    
    # Получаем подразделение
    dept_result = await db.execute(select(Department).where(Department.id == department_id))
    department = dept_result.scalar_one_or_none()
    
    if not department:
        raise HTTPException(status_code=404, detail="Подразделение не найдено")
    
    # Формируем запрос для получения нарядов
    query = (
        select(DutyRecord, Employee, DutyType)
        .join(Employee, DutyRecord.employee_id == Employee.id)
        .join(DutyType, DutyRecord.duty_type_id == DutyType.id)
        .where(Employee.department_id == department_id)
    )
    
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
        query = query.where(DutyRecord.duty_date >= start_dt).where(DutyRecord.duty_date <= end_dt)
    
    query = query.order_by(DutyRecord.duty_date, Employee.last_name, Employee.first_name)
    
    duty_records_result = await db.execute(query)
    duty_records = duty_records_result.all()
    
    # Формируем ответ
    distribution_data = []
    for duty_record, employee, duty_type in duty_records:
        distribution_data.append({
            'id': duty_record.id,
            'duty_type_name': duty_type.name,
            'date': duty_record.duty_date.isoformat(),
            'employee_name': f"{employee.last_name} {employee.first_name}",
            'department_name': department.name,
            'duty_count': employee.duty_count or 0  # Добавляем duty_count
        })
    
    return distribution_data 

@router.get("/duty-type/{duty_type_id}")
async def get_duty_distribution_by_type(
    duty_type_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Получить распределение нарядов по типу наряда (таблица: дата, ФИО, подразделение)"""
    duty_records_result = await db.execute(
        select(DutyRecord, Employee, Department)
        .join(Employee, DutyRecord.employee_id == Employee.id)
        .join(Department, Employee.department_id == Department.id)
        .where(DutyRecord.duty_type_id == duty_type_id)
        .order_by(DutyRecord.duty_date, Employee.last_name, Employee.first_name)
    )
    duty_records = duty_records_result.all()
    result = []
    for duty_record, employee, department in duty_records:
        result.append({
            'date': duty_record.duty_date.isoformat(),
            'employee_name': f"{employee.last_name} {employee.first_name}",
            'department_name': department.name
        })
    return result

@router.get("/all")
async def get_all_duties(
    year: int = Query(..., description="Год"),
    month: int = Query(..., description="Месяц"),
    db: AsyncSession = Depends(get_db)
):
    """Получить все наряды за месяц/год с группировкой по подразделениям"""
    
    # Получаем все наряды за месяц
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    duty_records_result = await db.execute(
        select(DutyRecord, Employee, Department, DutyType)
        .join(Employee, DutyRecord.employee_id == Employee.id)
        .join(Department, Employee.department_id == Department.id)
        .join(DutyType, DutyRecord.duty_type_id == DutyType.id)
        .where(DutyRecord.duty_date >= start_date)
        .where(DutyRecord.duty_date <= end_date)
        .order_by(DutyRecord.duty_date, Employee.last_name, Employee.first_name)
    )
    duty_records = duty_records_result.all()
    
    # Формируем ответ с дополнительной информацией о duty_count
    result = []
    for duty_record, employee, department, duty_type in duty_records:
        result.append({
            'id': duty_record.id,
            'date': duty_record.duty_date.isoformat(),
            'employee_id': employee.id,
            'employee_name': f"{employee.last_name} {employee.first_name}",
            'department_id': department.id,
            'department_name': department.name,
            'duty_type_id': duty_type.id,
            'duty_type_name': duty_type.name,
            'people_per_day': duty_type.people_per_day,
            'duty_count': employee.duty_count or 0  # Добавляем duty_count
        })
    
    return result

@router.delete("/clear")
async def clear_duty_records(
    request: DutyDistributionRequest,
    db: AsyncSession = Depends(get_db)
):
    """Удалить наряды за указанный период для конкретного подразделения"""
    from models.models import DutyRecord
    
    # Парсим даты
    start_date = datetime.strptime(request.start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(request.end_date, "%Y-%m-%d").date()
    
    # Формируем запрос для удаления нарядов
    delete_query = (
        select(DutyRecord)
        .where(DutyRecord.duty_date >= start_date)
        .where(DutyRecord.duty_date <= end_date)
    )
    
    # Если указано конкретное подразделение
    if request.department_id:
        delete_query = delete_query.where(Employee.department_id == request.department_id)
    # Если указана структура (но не подразделение)
    elif request.structure_id:
        # Получаем все подразделения структуры
        subdepts_result = await db.execute(
            select(Department.id).where(Department.parent_id == request.structure_id)
        )
        subdept_ids = [row[0] for row in subdepts_result.all()]
        if subdept_ids:
            delete_query = delete_query.where(Employee.department_id.in_(subdept_ids))
    
    result = await db.execute(delete_query)
    records = result.scalars().all()
    
    for record in records:
        await db.delete(record)
    await db.commit()
    
    return {"message": f"Удалено {len(records)} нарядов за период {start_date} - {end_date}"} 

@router.get("/export")
async def export_duties_to_excel(
    year: int = Query(..., description="Год"),
    month: int = Query(..., description="Месяц"),
    db: AsyncSession = Depends(get_db)
):
    """Экспортировать наряды за месяц/год в Excel (xlsx)"""
    try:
        # Получаем все записи нарядов за месяц/год
        duty_records_result = await db.execute(
            select(DutyRecord, Employee, Department, DutyType)
            .join(Employee, DutyRecord.employee_id == Employee.id)
            .join(Department, Employee.department_id == Department.id)
            .join(DutyType, DutyRecord.duty_type_id == DutyType.id)
            .where(func.extract('year', DutyRecord.duty_date) == year)
            .where(func.extract('month', DutyRecord.duty_date) == month)
            .order_by(DutyRecord.duty_date, Department.name, Employee.last_name, Employee.first_name)
        )
        duty_records = duty_records_result.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Наряды {month:02d}.{year}"
        ws.append(["Дата", "ФИО", "Подразделение", "Тип наряда"])
        if not duty_records:
            ws.append(["Нет данных", "", "", ""])
        else:
            for duty_record, employee, department, duty_type in duty_records:
                ws.append([
                    duty_record.duty_date.strftime("%d-%m-%Y"),
                    f"{employee.last_name} {employee.first_name}",
                    department.name,
                    duty_type.name
                ])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"duty_distribution_{year}_{month:02d}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export error: {e}") 

@router.get("/export/department/{department_id}")
async def export_department_duties_to_excel(
    department_id: int,
    year: int = Query(..., description="Год"),
    month: int = Query(..., description="Месяц"),
    format: str = Query("employees", description="Формат экспорта: employees (по сотрудникам) или duty_types (по датам и типам)"),
    db: AsyncSession = Depends(get_db)
):
    """Экспортировать наряды по подразделению в Excel"""
    try:
        from sqlalchemy import extract
        
        # Получаем все записи нарядов для подразделения за месяц/год
        duty_records_result = await db.execute(
            select(DutyRecord, Employee, DutyType)
            .join(Employee, DutyRecord.employee_id == Employee.id)
            .join(DutyType, DutyRecord.duty_type_id == DutyType.id)
            .where(Employee.department_id == department_id)
            .where(extract('year', DutyRecord.duty_date) == year)
            .where(extract('month', DutyRecord.duty_date) == month)
            .order_by(Employee.last_name, Employee.first_name, DutyRecord.duty_date)
        )
        duty_records = duty_records_result.all()
        
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if format == "duty_types":
            # Формат: даты в строках, типы нарядов в колонках
            ws.title = f"Подразделение {department_id} - по датам и типам"
            
            # Собираем уникальные даты и типы нарядов
            dates = sorted({duty_record.duty_date for duty_record, _, _ in duty_records})
            duty_types = []
            duty_type_set = set()
            for duty_record, _, duty_type in duty_records:
                if duty_type.id not in duty_type_set:
                    duty_types.append((duty_type.id, duty_type.name))
                    duty_type_set.add(duty_type.id)
            
            # Строим мапу: (date, duty_type_id) -> список сотрудников
            duty_map = {}
            for duty_record, employee, duty_type in duty_records:
                key = (duty_record.duty_date, duty_type.id)
                if key not in duty_map:
                    duty_map[key] = []
                duty_map[key].append(f"{employee.last_name} {employee.first_name}")
            
            # Функция получения цвета для типа наряда
            def get_duty_color(duty_type_name: str):
                name = duty_type_name.lower()
                if 'академический' in name:
                    return {'bg': 'E9D5FF', 'text': '7C3AED'}  # purple
                elif 'дежурный' in name or 'дежурство' in name:
                    return {'bg': 'DBEAFE', 'text': '1D4ED8'}  # blue
                elif 'патрульный' in name or 'патруль' in name:
                    return {'bg': 'D1FAE5', 'text': '047857'}  # green
                elif 'караульный' in name or 'караул' in name:
                    return {'bg': 'FEE2E2', 'text': 'DC2626'}  # red
                elif 'конвойный' in name or 'конвой' in name:
                    return {'bg': 'FED7AA', 'text': 'EA580C'}  # orange
                elif 'охранный' in name or 'охрана' in name:
                    return {'bg': 'FEF3C7', 'text': 'D97706'}  # yellow
                elif 'инспектор' in name:
                    return {'bg': 'E0E7FF', 'text': '3730A3'}  # indigo
                elif 'комендант' in name:
                    return {'bg': 'FCE7F3', 'text': 'BE185D'}  # pink
                else:
                    return {'bg': 'F3F4F6', 'text': '374151'}  # gray
            
            # Заголовки
            header = ["Дата"] + [duty_type_name for _, duty_type_name in duty_types]
            ws.append(header)
            
            # Применяем стили к заголовкам
            for col in range(1, len(header) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            
            # Данные по датам
            for row_idx, date in enumerate(dates, start=2):
                row = [date.strftime("%d-%m")]
                for col_idx, (duty_type_id, duty_type_name) in enumerate(duty_types, start=2):
                    employees = duty_map.get((date, duty_type_id), [])
                    cell_value = ", ".join(employees) if employees else "—"
                    row.append(cell_value)
                    
                    # Применяем цвет к ячейке если есть сотрудники
                    if employees:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        colors = get_duty_color(duty_type_name)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=colors['bg'], 
                            end_color=colors['bg'], 
                            fill_type="solid"
                        )
                        cell.font = openpyxl.styles.Font(
                            color=colors['text'],
                            bold=True
                        )
                        # Добавляем границы
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                
                ws.append(row)
                
            filename = f"department_{department_id}_duty_types_{year}_{month:02d}.xlsx"
        else:
            # Формат по умолчанию: сотрудники в строках, даты в колонках
            ws.title = f"Подразделение {department_id}"
            
            # Собираем уникальные даты и сотрудников
            dates = sorted({duty_record.duty_date for duty_record, _, _ in duty_records})
            employees = []
            emp_set = set()
            for duty_record, employee, _ in duty_records:
                if employee.id not in emp_set:
                    employees.append((employee.id, f"{employee.last_name} {employee.first_name}"))
                    emp_set.add(employee.id)
            
            # Строим мапу: (employee_id, date) -> тип наряда
            duty_map = {}
            for duty_record, employee, duty_type in duty_records:
                duty_map[(employee.id, duty_record.duty_date)] = duty_type.name
            
            # Первая строка — даты
            header = ["Сотрудник"] + [date.strftime("%d-%m") for date in dates]
            ws.append(header)
            
            # Данные по сотрудникам
            for emp_id, emp_name in employees:
                row = [emp_name]
                for date in dates:
                    row.append(duty_map.get((emp_id, date), ""))
                ws.append(row)
                
            filename = f"department_{department_id}_duties_{year}_{month:02d}.xlsx"
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export error: {e}") 